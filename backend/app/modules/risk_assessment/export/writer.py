from __future__ import annotations

from datetime import datetime
from io import BytesIO
from math import ceil
from unicodedata import east_asian_width

from app.core.datetime import BEIJING_TZ
from app.modules.risk_assessment.export.layout import (
    BORDER_COLOR,
    COLUMN_WIDTHS,
    DATA_LINE_HEIGHT,
    DATA_ROW_HEIGHT,
    FONT_NAME,
    HEADER_FILL,
    HEADER_FONT_COLOR,
    HEADER_ROW_HEIGHT,
    HEADERS,
    MERGED_RANGES,
    META_ROW_HEIGHT,
    MAX_DATA_ROW_HEIGHT,
    SECTION_FONT_COLOR,
    SECTION_TITLE,
    SHEET_NAME,
    TITLE,
    TITLE_ROW_HEIGHT,
)
from app.modules.risk_assessment.overview.schemas import BusinessOverviewProjection


class RiskAuditWorkbookWriter:
    def write(
        self,
        *,
        projection: BusinessOverviewProjection,
        compiled_at: datetime,
    ) -> BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        sheet.sheet_view.showGridLines = False
        for merged_range in MERGED_RANGES:
            sheet.merge_cells(merged_range)

        sheet["A1"] = TITLE
        sheet["A2"] = (
            f"业务编号: {projection.business_code}  |  "
            f"编制日期: {compiled_at.astimezone(BEIJING_TZ):%Y-%m-%d}"
        )
        sheet["A4"] = SECTION_TITLE
        for column, header in enumerate(HEADERS, start=1):
            sheet.cell(row=5, column=column, value=header)

        for row_index, row in enumerate(projection.rows, start=6):
            sheet.cell(row=row_index, column=1, value=row.label)
            sheet.cell(row=row_index, column=2, value=row.content)
            source_text = "；".join(row.source_files)
            if row.is_human_reviewed:
                source_text = f"{source_text}（人工复核）" if source_text else "人工复核"
            sheet.cell(row=row_index, column=3, value=source_text or "未识别")

        thin = Side(style="thin", color=BORDER_COLOR)
        data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet["A1"].font = Font(name=FONT_NAME, size=14, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A2"].font = Font(name=FONT_NAME, size=10, color="666666")
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A4"].font = Font(name=FONT_NAME, size=11, bold=True, color=SECTION_FONT_COLOR)
        sheet["A4"].alignment = Alignment(horizontal="left", vertical="center")

        for cell in sheet[5]:
            if cell.column > 3:
                continue
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=HEADER_FONT_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = data_border

        for row in sheet.iter_rows(min_row=6, max_row=22, min_col=1, max_col=3):
            for cell in row:
                cell.font = Font(name=FONT_NAME, size=10)
                cell.alignment = Alignment(
                    horizontal="left" if cell.column == 1 else "center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = data_border

        for column, width in COLUMN_WIDTHS.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[1].height = TITLE_ROW_HEIGHT
        sheet.row_dimensions[2].height = META_ROW_HEIGHT
        sheet.row_dimensions[5].height = HEADER_ROW_HEIGHT
        for row_index in range(6, 23):
            values = [sheet.cell(row=row_index, column=column).value for column in range(1, 4)]
            sheet.row_dimensions[row_index].height = _data_row_height(values)
        sheet.freeze_panes = "A6"
        sheet.print_area = "A1:C22"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        content = BytesIO()
        workbook.save(content)
        content.seek(0)
        return content


def _data_row_height(values: list[object | None]) -> float:
    wrapped_lines = max(
        _wrapped_line_count(value, COLUMN_WIDTHS[column])
        for column, value in zip(("A", "B", "C"), values, strict=True)
    )
    return min(MAX_DATA_ROW_HEIGHT, max(DATA_ROW_HEIGHT, wrapped_lines * DATA_LINE_HEIGHT))


def _wrapped_line_count(value: object | None, column_width: float) -> int:
    if value is None:
        return 1
    lines = str(value).splitlines() or [""]
    return sum(
        max(1, ceil(sum(_character_width(char) for char in line) / column_width))
        for line in lines
    )


def _character_width(character: str) -> int:
    return 2 if east_asian_width(character) in {"W", "F", "A"} else 1
